# -*- coding: utf-8 -*-
import pandas as pd
import os
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.views import SheetView
from datetime import datetime
import jdatetime

# TODO: use better method for last-night duplicate deletion
# TODO: speed up search.xlsx loading time

# تاریخ شمسی برای افزودن به نام فایل‌ها
today_jalali = jdatetime.date.today().strftime("%y%m%d")

# مسیر پوشه takhsis روی دسکتاپ هر کاربر
user_desktop = os.path.join(os.path.expanduser("~"), "Desktop", "takhsis")

in_wait_path = os.path.join(user_desktop, "in-wait.xlsx")
last_night_path = os.path.join(user_desktop, "last-night.xlsx")
search_path = os.path.join(user_desktop, "search.xlsx")
report_day_path = os.path.join(user_desktop, "takhsisReport.xlsx")
report_month_path = os.path.join(user_desktop, "takhsisReport-m.xlsx")
rating_path = os.path.join(user_desktop, "rating.xlsx")

# --- بارگذاری فایل‌ها ---
print("📂 Reading file in-wait in:", in_wait_path)
in_wait = pd.read_excel(in_wait_path,  dtype={"کد پذیرنده": str, "سریال پوز تخصیص یافته": str})

print("📂 Reading file last-night in:", last_night_path)
last_night = pd.read_excel(last_night_path, dtype={"کد پذیرنده": str})

print("📂 Reading file search in:", search_path)
search = pd.read_excel(search_path, dtype={"کد پذیرنده": str, "سریال پایانه": str})

# گزارش‌های تخصیص (ممکن است نباشند)
try:
    print("📂 Reading file takhsisReport in:", report_day_path)
    report_day = pd.read_excel(report_day_path)
except Exception:
    report_day = pd.DataFrame()

try:
    print("📂 Reading file takhsisReport-m in:", report_month_path)
    report_month = pd.read_excel(report_month_path)
except Exception:
    report_month = pd.DataFrame()

print("📂 Reading file rating in:", rating_path)
rating = pd.read_excel(rating_path, dtype={"کد پذیرنده": str})

# حذف ستون‌های اضافی از in-wait
columns_to_keep = [
    "کد پذیرنده", "کد درخواست", "کد پیگیری", "مدل پوز", "گروه پروژه",
    "تاریخ ایجاد", "آخرین تاریخ ویرایش", "شماره حساب"
]
in_wait = in_wait[[col for col in columns_to_keep if col in in_wait.columns]]

# Helper: POS type based on model
def get_pos_type(model):
    if pd.isna(model): return ""
    if str(model).strip().upper() == "GPRS":
        return "بیسیم"
    elif str(model).strip().upper() in ["LAN", "DIALUP", "PCPOSLAN"]:
        return "ثابت"
    return "نامشخص"

# Merge step 1: in-wait with search (on کد پذیرنده)
merged = pd.merge(
    in_wait,
    search[["کد پذیرنده", "نام فروشگاه", "شهر", "آدرس"]],
    on="کد پذیرنده",
    how="left"
)

# Merge step 2: with last-night for پشتیبان و توضیح (تنها یک ردیف به‌ازای هر کد پذیرنده)
last_night_info = last_night[["کد پذیرنده", "نام پشتیبان", "نام خانوادگی پشتیبان", "توضیح"]].copy()
last_night_info["نام و نام خانوادگی پشتیبان"] = last_night_info["نام پشتیبان"].fillna("") + " " + last_night_info["نام خانوادگی پشتیبان"].fillna("")
last_night_info = last_night_info.drop(columns=["نام پشتیبان", "نام خانوادگی پشتیبان"])
last_night_info = last_night_info.rename(columns={"توضیح": "توضیحات"})
last_night_info = last_night_info.drop_duplicates(subset="کد پذیرنده")

merged = pd.merge(
    merged,
    last_night_info,
    on="کد پذیرنده",
    how="left"
)

# Merge step 3: with rating for پله درآمد
merged = pd.merge(
    merged,
    rating,
    on="کد پذیرنده",
    how="left"
)

# توضیحات نهایی با در نظر گرفتن شرایط خاص
mask_missing = ~merged["کد پذیرنده"].isin(last_night_info["کد پذیرنده"])
merged.loc[mask_missing, "توضیحات"] = "نصب اولیه"
merged.loc[merged["توضیحات"] == "--", "توضیحات"] = ""
merged["توضیحات"] = merged["توضیحات"].fillna("")
merged["توضیحات"] = merged["توضیحات"].apply(lambda x: x.split(" - ")[0].strip() if isinstance(x, str) and " - " in x else x)

# گروه پایانه بر اساس مدل پوز
merged["گروه پایانه"] = merged["مدل پوز"].apply(get_pos_type)

# انتخاب و ساخت جدول نهایی
result_cols = [
    "کد پذیرنده", "نام فروشگاه", "پله درآمد", "توضیحات", "شهر", "آدرس", "نام و نام خانوادگی پشتیبان",
    "گروه پایانه", "کد درخواست", "کد پیگیری", "مدل پوز", "گروه پروژه",
    "تاریخ ایجاد", "آخرین تاریخ ویرایش", "شماره حساب"
]

final_result = merged[[col for col in result_cols if col in merged.columns]].copy()

# ساخت فایل نصب اولیه و فیلتر پروژه فروش
initial_installs = final_result[final_result["توضیحات"] == "نصب اولیه"].copy()
filtered_result = final_result[final_result["گروه پروژه"] != "پروژه فروش"].copy()

# ذخیره فایل تخصیص با تنظیم راست‌به‌چپ
output_path = os.path.join(user_desktop, f"takhsis{today_jalali}.xlsx")
filtered_result.to_excel(output_path, index=False, sheet_name="نتیجه")
wb = load_workbook(output_path)
ws = wb["نتیجه"]
ws.sheet_view.rightToLeft = True
wb.save(output_path)

# ذخیره فایل نصب اولیه با تنظیم راست‌به‌چپ
initial_path = os.path.join(user_desktop, f"نصب اولیه{today_jalali}.xlsx")
initial_installs.to_excel(initial_path, index=False, sheet_name="نصب اولیه")
wb2 = load_workbook(initial_path)
ws2 = wb2["نصب اولیه"]
ws2.sheet_view.rightToLeft = True
wb2.save(initial_path)

# ---- ساخت «گزارش تخصیص» با فرمت نمونه ----
# شمارش «در انتظار تخصیص»
waiting_total = len(filtered_result)
waiting_fixed = (filtered_result["گروه پایانه"] == "ثابت").sum()
waiting_wireless = (filtered_result["گروه پایانه"] == "بیسیم").sum()

# Helper: شمارش پروژه‌ها
def _count_projects(df: pd.DataFrame):
    if df is None or df.empty or ("پروژه" not in df.columns):
        return {"ps": 0, "sales": 0, "bank": 0, "total": 0}
    col = df["پروژه"].astype(str).fillna("").str.strip()
    ps = col.str.contains("پرشین|پرشين", case=False, regex=True).sum()
    sales = col.str.contains("فروش", case=False, regex=True).sum()
    total = len(df)
    bank = total - ps - sales
    return {"ps": int(ps), "sales": int(sales), "bank": int(bank), "total": int(total)}

cnt_day = _count_projects(report_day)
cnt_month = _count_projects(report_month)

wb_report = Workbook()
wsr = wb_report.active
wsr.title = "گزارش"

# ردیف‌ها مطابق فایل نمونه
wsr["A1"] = "گزارشات"; wsr["B1"] = "تعداد"
wsr["A2"] = "کل در انتظار تخصیص "; wsr["B2"] = waiting_total
wsr["A3"] = "در انتظار تخصیص ثابت"; wsr["B3"] = waiting_fixed
wsr["A4"] = "در انتظار تخصیص سیار"; wsr["B4"] = waiting_wireless
wsr["A5"] = "تعداد تخصیص پوز روز قبل پروژه بانکی "; wsr["B5"] = cnt_day["bank"]
wsr["A6"] = " تعداد تخصیص پوز روز قبل پرشین"; wsr["B6"] = cnt_day["ps"]
wsr["A7"] = "تعداد تخصیص پوز روز قبل پروژه فروش"; wsr["B7"] = cnt_day["sales"]
wsr["A8"] = "تعداد تخصیص پوز این ماه پروژه بانکی "; wsr["B8"] = cnt_month["bank"]
wsr["A9"] = " تعداد تخصیص پوز این ماه پرشین"; wsr["B9"] = cnt_month["ps"]
wsr["A10"] = "تعداد تخصیص پوز این ماه پروژه فروش"; wsr["B10"] = cnt_month["sales"]

# ادغام و جمع کل‌ها (همانند فایل نمونه)
wsr.merge_cells("C5:C7"); wsr.merge_cells("D5:D7")
wsr.merge_cells("C8:C10"); wsr.merge_cells("D8:D10")
wsr["D5"] = "جمع کل تخصیص یافته "; wsr["C5"] = cnt_day["total"]
wsr["D8"] = "جمع کل تخصیص یافته "; wsr["C8"] = cnt_month["total"]

# راست‌به‌چپ و عرض ستون‌ها
wsr.sheet_view.rightToLeft = True
from openpyxl.utils import get_column_letter
for col, width in zip(["A", "B", "C", "D"], [42, 12, 16, 24]):
    wsr.column_dimensions[col].width = width

allocation_path = os.path.join(user_desktop, f"گزارش تخصیص{today_jalali}.xlsx")
wb_report.save(allocation_path)

print("\n✅ فایل‌ها ذخیره شدند!\n📁", output_path, "\n📁", initial_path, "\n📁", allocation_path)
